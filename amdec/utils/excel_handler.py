"""
Module de gestion des imports/exports Excel pour l'application AMDEC.

Ce module fournit les classes ExcelImporter et ExcelExporter pour gérer
l'import de fichiers Excel AMDEC et l'export de projets au format Excel.

Standards:
    - Support UTF-8 complet (caractères français)
    - Validation des scores G/O/D (1-10)
    - Export avec mise en forme professionnelle
    - Gestion des logos
    - Calcul automatique de la criticité
"""

import io
import logging
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple, Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

from django.core.exceptions import ValidationError
from django.db import transaction
from django.contrib.auth.models import User

# Import local - sera disponible après création
# from ..models import AMDECProject, FailureMode, AMDECLogo

logger = logging.getLogger(__name__)


class ExcelImporter:
    """
    Classe pour importer des projets AMDEC depuis des fichiers Excel.

    Gère le parsing, la validation et la création de projets avec leurs
    défaillances associées à partir de fichiers Excel.
    """
    # Ajouter en haut de la classe
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

    def validate_file_size(self, file):
        """Valide la taille du fichier."""
        if file.size > self.MAX_FILE_SIZE:
            raise ValidationError(
                f"Le fichier dépasse la taille maximale "
                f"({self.MAX_FILE_SIZE / 1024 / 1024}MB)"
            )

    # Mapping des colonnes attendues
    COLUMN_MAPPING = {
        'Composant': 'component',
        'Mode de Défaillance': 'failure_mode',
        'Cause Potentielle': 'potential_cause',
        'Effet': 'effect',
        'Gravité (G)': 'gravity',
        'Occurrence (O)': 'occurrence',
        'Détection (D)': 'detection',
        'Actions Préventives': 'preventive_actions'
    }

    # Mots-clés pour détecter les métadonnées
    METADATA_KEYWORDS = {
        'système': 'system_name',
        'date': 'analysis_date',
        'équipe': 'team_members',
        'objectif': 'objective',
        'client': 'client',
        'référence': 'reference'
    }

    def __init__(self):
        """Initialise l'importeur avec les paramètres par défaut."""
        self.errors = []
        self.warnings = []
        self.imported_count = 0

    def parse_excel(self, file) -> Dict[str, Any]:
        """
        Parse un fichier Excel AMDEC et extrait les données.

        Args:
            file: Objet fichier Excel (UploadedFile ou BytesIO)

        Returns:
            dict: Dictionnaire contenant:
                - metadata: Informations du projet
                - failures: Liste des défaillances
                - summary: Statistiques du fichier

        Raises:
            ValidationError: Si le fichier est invalide ou corrompu
        """
        try:
            # Charger le workbook avec gestion UTF-8
            wb = load_workbook(file, data_only=True, read_only=False)
        except Exception as e:
            logger.error(f"Erreur lors du chargement du fichier Excel: {e}")
            raise ValidationError(f"Fichier Excel invalide: {str(e)}")

        # Vérifier la présence de la feuille principale
        if not wb.sheetnames:
            raise ValidationError("Le fichier Excel ne contient aucune feuille")

        # Prendre la première feuille ou celle nommée "Analyse AMDEC"
        ws = None
        for name in ['Analyse AMDEC', 'AMDEC', 'Analysis']:
            if name in wb.sheetnames:
                ws = wb[name]
                break

        if ws is None:
            ws = wb[wb.sheetnames[0]]

        # Extraire les données
        metadata = self._extract_metadata(ws)
        failures = self._extract_failures(ws)
        summary = self._calculate_summary(failures)

        # Nettoyer les ressources
        wb.close()

        return {
            'metadata': metadata,
            'failures': failures,
            'summary': summary
        }

    def _extract_metadata(self, worksheet) -> Dict[str, Any]:
        """
        Extrait les métadonnées du projet depuis la feuille Excel.

        Args:
            worksheet: Feuille Excel openpyxl

        Returns:
            dict: Métadonnées du projet
        """
        metadata = {
            'system_name': '',
            'analysis_date': date.today(),
            'team_members': '',
            'objective': '',
            'client': '',
            'reference': f"AMDEC-{datetime.now().strftime('%Y%m%d-%H%M')}"
        }

        # Parcourir les 15 premières lignes pour trouver les métadonnées
        for row in worksheet.iter_rows(min_row=1, max_row=15, values_only=True):
            if not row or not row[0]:
                continue

            # Convertir en string et nettoyer
            first_cell = str(row[0]).lower().strip()

            # Chercher les mots-clés
            for keyword, field in self.METADATA_KEYWORDS.items():
                if keyword in first_cell:
                    # Prendre la valeur de la cellule suivante ou de la même cellule
                    if len(row) > 1 and row[1]:
                        value = str(row[1]).strip()

                        # Traitement spécial pour les dates
                        if field == 'analysis_date':
                            metadata[field] = self._parse_date(value)
                        else:
                            metadata[field] = value
                    break

        # Extraction du nom depuis l'en-tête si présent
        for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and row[0]:
                cell_value = str(row[0])
                if 'STEP' in cell_value or 'AMDEC' in cell_value:
                    # Extraire le nom du système depuis l'en-tête
                    lines = cell_value.split('\n')
                    if lines:
                        # Nettoyer et extraire le nom
                        name = lines[0].replace('\r', '').strip()
                        if not metadata['system_name'] and name:
                            metadata['system_name'] = name
                    break

        return metadata

    def _extract_failures(self, worksheet) -> List[Dict[str, Any]]:
        """
        Extrait les modes de défaillance depuis la feuille Excel.

        Args:
            worksheet: Feuille Excel openpyxl

        Returns:
            list: Liste des défaillances avec leurs données
        """
        failures = []
        header_row = None
        column_indices = {}

        # Trouver la ligne d'en-têtes (contenant "Composant", "Mode", etc.)
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if not row:
                continue

            # Vérifier si c'est la ligne d'en-têtes
            row_str = ' '.join(str(cell) for cell in row if cell)
            if 'Composant' in row_str and ('Mode' in row_str or 'Défaillance' in row_str):
                header_row = row_idx

                # Mapper les indices de colonnes
                for col_idx, cell in enumerate(row):
                    if cell:
                        cell_str = str(cell).strip()
                        for excel_name, field_name in self.COLUMN_MAPPING.items():
                            if excel_name in cell_str or excel_name.lower() in cell_str.lower():
                                column_indices[field_name] = col_idx
                                break
                break

        if not header_row or not column_indices:
            self.warnings.append("Impossible de trouver les en-têtes de colonnes")
            return failures

        # Extraire les données de défaillance
        for row_idx, row in enumerate(worksheet.iter_rows(
                min_row=header_row + 1, values_only=True), start=header_row + 1):

            if not row or not any(row):
                continue

            # Vérifier si c'est une ligne de données (pas de résumé)
            first_cell = str(row[0]) if row[0] else ''
            if any(keyword in first_cell.lower() for keyword in
                   ['résumé', 'total', 'criticité élevée', 'criticité modérée']):
                break

            # Extraire les données de la ligne
            failure_data = {
                'order': len(failures)
            }

            for field_name, col_idx in column_indices.items():
                if col_idx < len(row):
                    value = row[col_idx]

                    # Nettoyer et convertir les valeurs
                    if field_name in ['gravity', 'occurrence', 'detection']:
                        failure_data[field_name] = self._parse_score(value, field_name, row_idx)
                    else:
                        failure_data[field_name] = self._clean_text(value) if value else ''

            # Valider que les champs requis sont présents
            if failure_data.get('component') and failure_data.get('failure_mode'):
                failures.append(failure_data)
            else:
                self.warnings.append(f"Ligne {row_idx}: Données incomplètes ignorées")

        return failures

    def _parse_date(self, value: Any) -> date:
        """
        Parse une date depuis différents formats.

        Args:
            value: Valeur à parser (string, date, datetime)

        Returns:
            date: Date parsée
        """
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, str):
            # Essayer différents formats
            formats = [
                '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y',
                '%Y/%m/%d', '%d.%m.%Y'
            ]

            for fmt in formats:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except ValueError:
                    continue

        # Par défaut, retourner aujourd'hui
        self.warnings.append(f"Date invalide '{value}', utilisation de la date du jour")
        return date.today()

    def _parse_score(self, value: Any, field_name: str, row_num: int) -> int:
        """
        Parse et valide un score (G/O/D).

        Args:
            value: Valeur à parser
            field_name: Nom du champ (pour les messages d'erreur)
            row_num: Numéro de ligne (pour les messages d'erreur)

        Returns:
            int: Score entre 1 et 10
        """
        try:
            score = int(float(str(value)))
            if 1 <= score <= 10:
                return score
            else:
                self.warnings.append(
                    f"Ligne {row_num}: {field_name} = {score} hors limites [1-10], "
                    f"valeur limitée à {min(10, max(1, score))}"
                )
                return min(10, max(1, score))
        except (ValueError, TypeError):
            self.errors.append(
                f"Ligne {row_num}: {field_name} invalide '{value}', "
                f"utilisation de la valeur par défaut 5"
            )
            return 5

    def _clean_text(self, value: Any) -> str:
        """
        Nettoie et normalise du texte.

        Args:
            value: Valeur à nettoyer

        Returns:
            str: Texte nettoyé
        """
        if value is None:
            return ''

        # Convertir en string et nettoyer
        text = str(value).strip()

        # Remplacer les retours à la ligne multiples
        text = text.replace('\r\n', '\n').replace('\r', '\n')

        # Supprimer les espaces multiples
        text = ' '.join(text.split())

        return text

    def _calculate_summary(self, failures: List[Dict]) -> Dict[str, int]:
        """
        Calcule les statistiques du projet.

        Args:
            failures: Liste des défaillances

        Returns:
            dict: Statistiques (total, high, medium, low)
        """
        summary = {
            'total': len(failures),
            'high': 0,
            'medium': 0,
            'low': 0
        }

        for failure in failures:
            # Calculer la criticité si les scores sont présents
            if all(k in failure for k in ['gravity', 'occurrence', 'detection']):
                criticality = (
                        failure['gravity'] *
                        failure['occurrence'] *
                        failure['detection']
                )

                if criticality > 120:
                    summary['high'] += 1
                elif criticality > 60:
                    summary['medium'] += 1
                else:
                    summary['low'] += 1

        return summary

    def validate_data(self, data: Dict[str, Any]) -> bool:
        """
        Valide les données extraites avant import.

        Args:
            data: Données à valider

        Returns:
            bool: True si valide, False sinon
        """
        self.errors.clear()

        # Vérifier les métadonnées requises
        metadata = data.get('metadata', {})
        if not metadata.get('system_name'):
            self.errors.append("Nom du système manquant")

        # Vérifier les défaillances
        failures = data.get('failures', [])
        if not failures:
            self.errors.append("Aucune défaillance trouvée dans le fichier")
            return False

        # Valider chaque défaillance
        for idx, failure in enumerate(failures, 1):
            if not failure.get('component'):
                self.errors.append(f"Défaillance {idx}: Composant manquant")

            if not failure.get('failure_mode'):
                self.errors.append(f"Défaillance {idx}: Mode de défaillance manquant")

            # Vérifier les scores
            for field in ['gravity', 'occurrence', 'detection']:
                value = failure.get(field)
                if value is None:
                    self.errors.append(f"Défaillance {idx}: {field} manquant")
                elif not isinstance(value, int) or not 1 <= value <= 10:
                    self.errors.append(f"Défaillance {idx}: {field}={value} invalide (doit être 1-10)")

        return len(self.errors) == 0

    @transaction.atomic
    def create_project(self, data: Dict[str, Any], user: User) -> 'AMDECProject':
        """
        Crée un projet AMDEC avec ses défaillances.

        Args:
            data: Données validées du projet
            user: Utilisateur créateur

        Returns:
            AMDECProject: Projet créé

        Raises:
            ValidationError: Si la création échoue
        """
        from ..models import AMDECProject, FailureMode

        # Valider les données
        if not self.validate_data(data):
            raise ValidationError(
                f"Données invalides: {'; '.join(self.errors)}"
            )

        metadata = data['metadata']
        failures = data['failures']

        try:
            # Créer le projet
            project = AMDECProject.objects.create(
                name=metadata.get('system_name', 'Projet importé'),
                reference=metadata.get('reference'),
                client=metadata.get('client', 'Non spécifié'),
                analysis_date=metadata.get('analysis_date', date.today()),
                team_members=metadata.get('team_members', ''),
                objective=metadata.get('objective', ''),
                description=f"Importé depuis Excel le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
                created_by=user,
                modified_by=user
            )

            # Créer les défaillances
            for failure_data in failures:
                FailureMode.objects.create(
                    project=project,
                    component=failure_data.get('component', ''),
                    failure_mode=failure_data.get('failure_mode', ''),
                    potential_cause=failure_data.get('potential_cause', ''),
                    effect=failure_data.get('effect', ''),
                    gravity=failure_data.get('gravity', 5),
                    occurrence=failure_data.get('occurrence', 5),
                    detection=failure_data.get('detection', 5),
                    preventive_actions=failure_data.get('preventive_actions', ''),
                    order=failure_data.get('order', 0),
                    status='PENDING'
                )
                self.imported_count += 1

            logger.info(
                f"Projet {project.reference} créé avec {self.imported_count} défaillances "
                f"par {user.username}"
            )

            return project

        except Exception as e:
            logger.error(f"Erreur lors de la création du projet: {e}")
            raise ValidationError(f"Erreur lors de la création: {str(e)}")


class ExcelExporter:
    """
    Classe pour exporter des projets AMDEC au format Excel.

    Génère des fichiers Excel professionnels avec mise en forme,
    calculs automatiques et support des logos.
    """

    # Styles prédéfinis
    STYLES = {
        'title': {
            'font': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='2C3E50', end_color='34495E', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        },
        'header': {
            'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='3498DB', end_color='2980B9', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'cell': {
            'font': Font(name='Calibri', size=10),
            'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )
        },
        'criticality_high': {
            'fill': PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'),
            'font': Font(color='FFFFFF', bold=True)
        },
        'criticality_medium': {
            'fill': PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid'),
            'font': Font(color='FFFFFF', bold=True)
        },
        'criticality_low': {
            'fill': PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'),
            'font': Font(color='FFFFFF', bold=True)
        }
    }

    def __init__(self):
        """Initialise l'exporteur avec une nouvelle workbook."""
        self.wb = Workbook()
        # Supprimer la feuille par défaut
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

    def export_project(self, project_id: int) -> io.BytesIO:
        """
        Exporte un projet AMDEC complet au format Excel.

        Args:
            project_id: ID du projet à exporter

        Returns:
            BytesIO: Fichier Excel en mémoire

        Raises:
            AMDECProject.DoesNotExist: Si le projet n'existe pas
        """
        from ..models import AMDECProject

        # Récupérer le projet avec ses relations
        project = AMDECProject.objects.select_related(
            'created_by', 'modified_by', 'logo'
        ).prefetch_related('failures').get(pk=project_id)

        # Générer les feuilles
        self.generate_summary_sheet(self.wb, project)
        self.generate_failures_sheet(self.wb, project)
        self.generate_actions_sheet(self.wb, project)

        # Ajouter la clause de non-responsabilité
        self._add_disclaimer_sheet()

        # Sauvegarder dans BytesIO
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)

        return output

    def generate_summary_sheet(self, wb: Workbook, project: 'AMDECProject'):
        """
        Génère la feuille de résumé du projet.

        Args:
            wb: Workbook openpyxl
            project: Projet AMDEC
        """
        ws = wb.create_sheet('Résumé', 0)

        # Ajouter le logo si disponible
        if project.logo and project.logo.image:
            try:
                img = XLImage(project.logo.image.path)
                img.width = 150
                img.height = 100
                ws.add_image(img, 'A1')
            except Exception as e:
                logger.warning(f"Impossible d'ajouter le logo: {e}")

        # Titre principal
        ws.merge_cells('A4:J4')
        title_cell = ws['A4']
        title_cell.value = f"{project.reference} - {project.name}"
        title_cell.font = self.STYLES['title']['font']
        title_cell.fill = self.STYLES['title']['fill']
        title_cell.alignment = self.STYLES['title']['alignment']
        ws.row_dimensions[4].height = 40

        # Sous-titre
        ws.merge_cells('A5:J5')
        ws['A5'] = "ANALYSE DES RISQUES DE DÉFAILLANCES (AMDEC)"
        ws['A5'].font = Font(size=14, bold=True)
        ws['A5'].alignment = Alignment(horizontal='center')

        # Informations du projet
        info_row = 7
        project_info = [
            ('Système analysé:', project.name),
            ('Client:', project.client),
            ('Date d\'analyse:', project.analysis_date.strftime('%d/%m/%Y')),
            ('Équipe AMDEC:', project.team_members),
            ('Objectif:', project.objective),
            ('Créé par:', project.created_by.get_full_name() or project.created_by.username),
            ('Date de création:', project.created_at.strftime('%d/%m/%Y %H:%M'))
        ]

        for label, value in project_info:
            ws[f'B{info_row}'] = label
            ws[f'B{info_row}'].font = Font(bold=True)
            ws.merge_cells(f'D{info_row}:I{info_row}')
            ws[f'D{info_row}'] = value
            info_row += 1

        # Échelle d'évaluation
        info_row += 1
        ws[f'B{info_row}'] = "ÉCHELLE D'ÉVALUATION"
        ws[f'B{info_row}'].font = Font(size=12, bold=True)

        info_row += 1
        scales = [
            "Gravité (G): 1-3 Mineure, 4-6 Modérée, 7-10 Critique",
            "Occurrence (O): 1-3 Rare, 4-6 Occasionnelle, 7-10 Fréquente",
            "Détection (D): 1-3 Facile, 4-6 Modérée, 7-10 Difficile",
            "Criticité (C): G × O × D | 1-60 Faible, 61-120 Modérée, 121-1000 Élevée"
        ]

        for scale in scales:
            ws.merge_cells(f'B{info_row}:I{info_row}')
            ws[f'B{info_row}'] = scale
            info_row += 1

        # Statistiques
        info_row += 2
        ws[f'B{info_row}'] = "STATISTIQUES"
        ws[f'B{info_row}'].font = Font(size=12, bold=True)

        info_row += 1
        stats = [
            ('Total des défaillances:', project.total_failures),
            ('Criticité élevée (>120):', project.high_criticality_count),
            ('Criticité modérée (61-120):', project.medium_criticality_count),
            ('Criticité faible (1-60):', project.low_criticality_count)
        ]

        for label, value in stats:
            ws[f'C{info_row}'] = label
            ws[f'C{info_row}'].font = Font(bold=True)
            ws[f'E{info_row}'] = value
            ws[f'E{info_row}'].font = Font(size=14, bold=True)

            # Colorer selon le niveau
            if 'élevée' in label.lower():
                ws[f'E{info_row}'].fill = self.STYLES['criticality_high']['fill']
                ws[f'E{info_row}'].font = self.STYLES['criticality_high']['font']
            elif 'modérée' in label.lower():
                ws[f'E{info_row}'].fill = self.STYLES['criticality_medium']['fill']
                ws[f'E{info_row}'].font = self.STYLES['criticality_medium']['font']
            elif 'faible' in label.lower():
                ws[f'E{info_row}'].fill = self.STYLES['criticality_low']['fill']
                ws[f'E{info_row}'].font = self.STYLES['criticality_low']['font']

            info_row += 1

        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 30
        for col in range(5, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def generate_failures_sheet(self, wb: Workbook, project: 'AMDECProject'):
        """
        Génère la feuille des modes de défaillance.

        Args:
            wb: Workbook openpyxl
            project: Projet AMDEC
        """
        ws = wb.create_sheet('Analyse AMDEC')

        # Titre
        ws.merge_cells('A1:J1')
        title = f"{project.reference} - {project.name}\nANALYSE DES MODES DE DÉFAILLANCE"
        ws['A1'] = title
        ws['A1'].font = self.STYLES['title']['font']
        ws['A1'].fill = self.STYLES['title']['fill']
        ws['A1'].alignment = self.STYLES['title']['alignment']
        ws.row_dimensions[1].height = 60

        # En-têtes du tableau
        headers = [
            'Composant',
            'Mode de Défaillance',
            'Cause Potentielle',
            'Effet',
            'Gravité (G)',
            'Occurrence (O)',
            'Détection (D)',
            'Criticité (C)',
            'Niveau de Criticité',
            'Actions Préventives'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.STYLES['header']['font']
            cell.fill = self.STYLES['header']['fill']
            cell.alignment = self.STYLES['header']['alignment']
            cell.border = self.STYLES['header']['border']

        # Données des défaillances
        failures = project.failures.all().order_by('-gravity', '-occurrence', '-detection')
        row_num = 4

        for failure in failures:
            # Données de la défaillance
            row_data = [
                failure.component,
                failure.failure_mode,
                failure.potential_cause,
                failure.effect,
                failure.gravity,
                failure.occurrence,
                failure.detection,
                failure.criticality,
                failure.criticality_level,
                failure.preventive_actions or ''
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = value
                cell.font = self.STYLES['cell']['font']
                cell.alignment = self.STYLES['cell']['alignment']
                cell.border = self.STYLES['cell']['border']

                # Coloration spéciale pour la criticité
                if col == 9:  # Colonne "Niveau de Criticité"
                    if failure.criticality_level == "ÉLEVÉE":
                        cell.fill = self.STYLES['criticality_high']['fill']
                        cell.font = self.STYLES['criticality_high']['font']
                    elif failure.criticality_level == "MODÉRÉE":
                        cell.fill = self.STYLES['criticality_medium']['fill']
                        cell.font = self.STYLES['criticality_medium']['font']
                    else:
                        cell.fill = self.STYLES['criticality_low']['fill']
                        cell.font = self.STYLES['criticality_low']['font']

                # Centrer les scores
                if col in [5, 6, 7, 8]:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            row_num += 1

        # Appliquer la mise en forme conditionnelle
        self.apply_conditional_formatting(ws)

        # Ajuster les largeurs de colonnes
        column_widths = [25, 30, 35, 35, 12, 15, 15, 15, 20, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Figer les volets (en-têtes)
        ws.freeze_panes = 'A4'

        # Ajouter des validations de données pour les scores
        for col_letter in ['E', 'F', 'G']:  # G, O, D
            dv = DataValidation(
                type="whole",
                operator="between",
                formula1=1,
                formula2=10,
                showDropDown=False
            )
            dv.error = 'La valeur doit être entre 1 et 10'
            dv.errorTitle = 'Valeur invalide'
            ws.add_data_validation(dv)
            dv.add(f'{col_letter}4:{col_letter}1000')

    def generate_actions_sheet(self, wb: Workbook, project: 'AMDECProject'):
        """
        Génère la feuille du plan d'actions prioritaires.

        Args:
            wb: Workbook openpyxl
            project: Projet AMDEC
        """
        ws = wb.create_sheet('Plan d\'Actions')

        # Titre
        ws.merge_cells('A1:F1')
        ws['A1'] = "PLAN D'ACTIONS PRIORITAIRES"
        ws['A1'].font = self.STYLES['title']['font']
        ws['A1'].fill = self.STYLES['title']['fill']
        ws['A1'].alignment = self.STYLES['title']['alignment']
        ws.row_dimensions[1].height = 40

        # Actions par niveau de criticité
        row_num = 3

        # Actions immédiates (Criticité élevée)
        high_failures = [f for f in project.failures.all() if f.criticality > 120]

        ws.merge_cells(f'A{row_num}:F{row_num}')
        ws[f'A{row_num}'] = "ACTIONS IMMÉDIATES (Criticité > 120)"
        ws[f'A{row_num}'].font = Font(size=12, bold=True, color='FFFFFF')
        ws[f'A{row_num}'].fill = self.STYLES['criticality_high']['fill']
        row_num += 1

        if high_failures:
            # En-têtes
            headers = ['Composant', 'Défaillance', 'Criticité', 'Actions Recommandées', 'Responsable', 'Échéance']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.border = self.STYLES['header']['border']
            row_num += 1

            # Données
            for failure in high_failures:
                ws.cell(row=row_num, column=1, value=failure.component)
                ws.cell(row=row_num, column=2, value=failure.failure_mode)
                ws.cell(row=row_num, column=3, value=failure.criticality)
                ws.cell(row=row_num, column=4, value=failure.preventive_actions or 'À définir')
                ws.cell(row=row_num, column=5, value=failure.responsible or 'À assigner')
                ws.cell(row=row_num, column=6,
                        value=failure.deadline.strftime('%d/%m/%Y') if failure.deadline else 'À planifier')

                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = self.STYLES['cell']['border']

                row_num += 1
        else:
            ws.merge_cells(f'A{row_num}:F{row_num}')
            ws[f'A{row_num}'] = "Aucune défaillance de criticité élevée"
            row_num += 1

        row_num += 2

        # Actions à moyen terme (Criticité modérée)
        medium_failures = [f for f in project.failures.all() if 61 <= f.criticality <= 120]

        ws.merge_cells(f'A{row_num}:F{row_num}')
        ws[f'A{row_num}'] = "ACTIONS À MOYEN TERME (Criticité 61-120)"
        ws[f'A{row_num}'].font = Font(size=12, bold=True, color='FFFFFF')
        ws[f'A{row_num}'].fill = self.STYLES['criticality_medium']['fill']
        row_num += 1

        if medium_failures:
            # En-têtes
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.border = self.STYLES['header']['border']
            row_num += 1

            # Données
            for failure in medium_failures:
                ws.cell(row=row_num, column=1, value=failure.component)
                ws.cell(row=row_num, column=5, value=failure.responsible or 'À assigner')
                ws.cell(row=row_num, column=6,
                        value=failure.deadline.strftime('%d/%m/%Y') if failure.deadline else 'À planifier')

                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).border = self.STYLES['cell']['border']

                row_num += 1
        else:
            ws.merge_cells(f'A{row_num}:F{row_num}')
            ws[f'A{row_num}'] = "Aucune défaillance de criticité modérée"
            row_num += 1

        # Ajuster les largeurs de colonnes
        column_widths = [25, 30, 12, 40, 20, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def apply_conditional_formatting(self, ws):
        """
        Applique la mise en forme conditionnelle à une feuille.

        Args:
            ws: Worksheet openpyxl
        """
        # Coloration de la colonne Criticité (H) selon la valeur
        # Rouge pour > 120
        red_fill = PatternFill(start_color='FFE74C3C', end_color='FFE74C3C', fill_type='solid')
        red_font = Font(color='FFFFFFFF', bold=True)
        ws.conditional_formatting.add(
            'H4:H1000',
            CellIsRule(operator='greaterThan', formula=['120'], fill=red_fill, font=red_font)
        )

        # Orange pour 61-120
        orange_fill = PatternFill(start_color='FFF39C12', end_color='FFF39C12', fill_type='solid')
        orange_font = Font(color='FFFFFFFF', bold=True)
        ws.conditional_formatting.add(
            'H4:H1000',
            CellIsRule(operator='between', formula=['61', '120'], fill=orange_fill, font=orange_font)
        )

        # Vert pour <= 60
        green_fill = PatternFill(start_color='FF27AE60', end_color='FF27AE60', fill_type='solid')
        green_font = Font(color='FFFFFFFF', bold=True)
        ws.conditional_formatting.add(
            'H4:H1000',
            CellIsRule(operator='lessThanOrEqual', formula=['60'], fill=green_fill, font=green_font)
        )

    def _add_disclaimer_sheet(self):
        """Ajoute une feuille avec la clause de non-responsabilité."""
        ws = self.wb.create_sheet('Mentions légales')

        ws.merge_cells('B2:H2')
        ws['B2'] = "CLAUSE DE NON-RESPONSABILITÉ"
        ws['B2'].font = Font(size=14, bold=True)
        ws['B2'].alignment = Alignment(horizontal='center')

        disclaimer_text = """
Cette analyse AMDEC est fournie à titre informatif uniquement et constitue une évaluation 
préliminaire des risques potentiels. Elle ne remplace pas une expertise professionnelle 
approfondie et ne doit pas être utilisée comme unique base de décision pour des questions 
de sécurité critiques.

L'utilisateur assume l'entière responsabilité de l'interprétation et de l'utilisation 
des informations contenues dans ce document. Les auteurs et l'organisation déclinent 
toute responsabilité pour les dommages directs ou indirects résultant de l'utilisation 
de cette analyse.

Il est fortement recommandé de faire valider cette analyse par des experts qualifiés 
dans le domaine concerné avant toute mise en œuvre des actions proposées.

Document généré automatiquement le {}.
        """.format(datetime.now().strftime('%d/%m/%Y à %H:%M'))

        ws.merge_cells('B4:H20')
        ws['B4'] = disclaimer_text.strip()
        ws['B4'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws['B4'].font = Font(size=10)

        # Ajuster les dimensions
        ws.column_dimensions['B'].width = 15
        for col in range(3, 9):
            ws.column_dimensions[get_column_letter(col)].width = 12
        ws.row_dimensions[4].height = 200
