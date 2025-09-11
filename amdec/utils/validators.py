"""
Module de validation pour l'application AMDEC.

Ce module fournit des fonctions de validation réutilisables pour
valider les données d'entrée, nettoyer le texte et parser les dates.

Standards:
    - Support UTF-8 complet (caractères français)
    - Messages d'erreur clairs et actionnables
    - Validation stricte des scores AMDEC
    - Gestion robuste des formats de date français
"""

import re
import unicodedata
from datetime import datetime, date
from typing import Any, Optional, Tuple, List, Dict
import logging
import magic

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from openpyxl import Workbook

logger = logging.getLogger(__name__)


# ============================================================================
# VALIDATION DES SCORES AMDEC
# ============================================================================

def validate_score(value: Any, field_name: str) -> int:
    """
    Valide qu'un score AMDEC est entre 1 et 10.
    
    Args:
        value: Valeur à valider (peut être int, float, str)
        field_name: Nom du champ pour les messages d'erreur
        
    Returns:
        int: Score validé entre 1 et 10
        
    Raises:
        ValidationError: Si la valeur n'est pas valide
    """
    # Conversion en entier
    try:
        if isinstance(value, str):
            # Nettoyer les espaces et virgules (format français)
            value = value.strip().replace(',', '.')
        
        score = int(float(value))
    except (ValueError, TypeError):
        raise ValidationError(
            f"❌ {field_name}: La valeur '{value}' n'est pas un nombre valide. "
            f"Utilisez un nombre entier entre 1 et 10."
        )
    
    # Validation de la plage
    if not 1 <= score <= 10:
        raise ValidationError(
            f"❌ {field_name}: La valeur {score} est hors limites. "
            f"Les scores doivent être entre 1 (minimum) et 10 (maximum)."
        )
    
    return score


def validate_criticality_scores(gravity: Any, occurrence: Any, detection: Any) -> Tuple[int, int, int]:
    """
    Valide un ensemble complet de scores de criticité.
    
    Args:
        gravity: Score de gravité
        occurrence: Score d'occurrence
        detection: Score de détection
        
    Returns:
        tuple: (gravity, occurrence, detection) validés
        
    Raises:
        ValidationError: Si au moins un score est invalide
    """
    errors = []
    
    try:
        gravity = validate_score(gravity, "Gravité (G)")
    except ValidationError as e:
        errors.append(str(e))
    
    try:
        occurrence = validate_score(occurrence, "Occurrence (O)")
    except ValidationError as e:
        errors.append(str(e))
    
    try:
        detection = validate_score(detection, "Détection (D)")
    except ValidationError as e:
        errors.append(str(e))
    
    if errors:
        raise ValidationError(errors)
    
    # Alerte pour criticité très élevée
    criticality = gravity * occurrence * detection
    if criticality > 500:
        logger.warning(
            f"⚠️ Criticité très élevée détectée: {criticality} "
            f"(G={gravity}, O={occurrence}, D={detection})"
        )
    
    return gravity, occurrence, detection


# ============================================================================
# VALIDATION DE FICHIERS EXCEL
# ============================================================================

def validate_excel_structure(wb: Workbook) -> Dict[str, Any]:
    """
    Vérifie qu'un workbook Excel a la structure attendue pour une AMDEC.
    
    Args:
        wb: Workbook openpyxl à valider
        
    Returns:
        dict: Informations sur la structure trouvée
            - has_analysis_sheet: bool
            - has_required_columns: bool
            - found_sheets: list
            - found_columns: list
            - warnings: list
            
    Raises:
        ValidationError: Si la structure est invalide
    """
    result = {
        'has_analysis_sheet': False,
        'has_required_columns': False,
        'found_sheets': [],
        'found_columns': [],
        'warnings': []
    }
    
    # Vérifier la présence de feuilles
    if not wb.sheetnames:
        raise ValidationError("❌ Le fichier Excel ne contient aucune feuille")
    
    result['found_sheets'] = wb.sheetnames
    
    # Chercher une feuille d'analyse
    analysis_sheets = ['Analyse AMDEC', 'AMDEC', 'Analysis', 'Analyse']
    analysis_sheet = None
    
    for sheet_name in wb.sheetnames:
        if any(keyword in sheet_name for keyword in analysis_sheets):
            analysis_sheet = wb[sheet_name]
            result['has_analysis_sheet'] = True
            break
    
    if not analysis_sheet:
        # Prendre la première feuille par défaut
        analysis_sheet = wb[wb.sheetnames[0]]
        result['warnings'].append(
            f"⚠️ Aucune feuille 'Analyse AMDEC' trouvée, utilisation de '{wb.sheetnames[0]}'"
        )
    
    # Vérifier les colonnes requises
    required_columns = [
        'Composant', 'Mode', 'Défaillance', 'Cause', 'Effet',
        'Gravité', 'Occurrence', 'Détection'
    ]
    
    # Parcourir les 20 premières lignes pour trouver les en-têtes
    for row in analysis_sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        if not row:
            continue
        
        row_text = ' '.join(str(cell) for cell in row if cell)
        found_cols = []
        
        for col in required_columns:
            if col.lower() in row_text.lower():
                found_cols.append(col)
        
        if len(found_cols) >= 5:  # Au moins 5 colonnes requises trouvées
            result['has_required_columns'] = True
            result['found_columns'] = found_cols
            break
    
    if not result['has_required_columns']:
        result['warnings'].append(
            "⚠️ Structure de colonnes AMDEC non trouvée. "
            "Colonnes attendues: Composant, Mode de Défaillance, Cause, Effet, G, O, D"
        )
    
    return result


def validate_excel_file(file) -> None:
    """
    Valide un fichier Excel uploadé pour l'import AMDEC.
    
    Args:
        file: Fichier uploadé (Django UploadedFile)
        
    Raises:
        ValidationError: Si le fichier n'est pas valide
    """
    # Vérifier l'extension
    if not hasattr(file, 'name'):
        raise ValidationError("❌ Fichier invalide: nom de fichier manquant")
    
    if not file.name.endswith(('.xlsx', '.xls')):
        raise ValidationError(
            f"❌ Format non supporté: '{file.name}'. "
            f"Utilisez un fichier Excel (.xlsx ou .xls)"
        )
    
    # Vérifier la taille (10MB max)
    if file.size > 10 * 1024 * 1024:
        size_mb = file.size / (1024 * 1024)
        raise ValidationError(
            f"❌ Fichier trop volumineux: {size_mb:.1f}MB. "
            f"La taille maximale autorisée est de 10MB"
        )
    
    # Vérifier le contenu avec python-magic si disponible
    try:
        file_type = magic.from_buffer(file.read(1024), mime=True)
        file.seek(0)  # Reset file pointer
        
        valid_types = [
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/octet-stream'  # Parfois retourné pour xlsx
        ]
        
        if file_type not in valid_types:
            logger.warning(f"Type MIME suspect: {file_type}")
            raise ValidationError(
                f"❌ Le fichier ne semble pas être un vrai fichier Excel. "
                f"Type détecté: {file_type}"
            )
    except Exception as e:
        # Si python-magic n'est pas disponible, on continue
        logger.debug(f"Validation MIME ignorée: {e}")


# ============================================================================
# NETTOYAGE ET NORMALISATION DE TEXTE
# ============================================================================

def clean_text(text: Any, max_length: Optional[int] = None, 
               preserve_newlines: bool = True) -> str:
    """
    Nettoie et normalise du texte en préservant les caractères français.
    
    Args:
        text: Texte à nettoyer (peut être None, str, ou autre)
        max_length: Longueur maximale (tronque si dépassée)
        preserve_newlines: Si True, préserve les sauts de ligne
        
    Returns:
        str: Texte nettoyé et normalisé
    """
    if text is None:
        return ''
    
    # Convertir en string
    text = str(text).strip()
    
    if not text:
        return ''
    
    # Normaliser les caractères Unicode (NFC pour le français)
    text = unicodedata.normalize('NFC', text)
    
    # Remplacer les caractères de contrôle sauf newline
    if preserve_newlines:
        # Garder uniquement \n, remplacer \r\n et \r
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        # Supprimer les autres caractères de contrôle
        text = ''.join(char for char in text if ord(char) >= 32 or char == '\n')
    else:
        # Remplacer tous les sauts de ligne par des espaces
        text = text.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
        # Supprimer tous les caractères de contrôle
        text = ''.join(char for char in text if ord(char) >= 32)
    
    # Supprimer les espaces multiples
    text = re.sub(r' +', ' ', text)
    
    # Supprimer les sauts de ligne multiples si préservés
    if preserve_newlines:
        text = re.sub(r'\n{3,}', '\n\n', text)
    
    # Nettoyer les espaces autour des sauts de ligne
    if preserve_newlines:
        lines = text.split('\n')
        lines = [line.strip() for line in lines]
        text = '\n'.join(lines)
    
    # Tronquer si nécessaire
    if max_length and len(text) > max_length:
        text = text[:max_length-3] + '...'
    
    return text.strip()


def sanitize_filename(filename: str) -> str:
    """
    Nettoie un nom de fichier pour le rendre sûr.
    
    Args:
        filename: Nom de fichier à nettoyer
        
    Returns:
        str: Nom de fichier sûr
    """
    # Garder seulement l'extension
    if '.' in filename:
        name, ext = filename.rsplit('.', 1)
        ext = '.' + ext.lower()
    else:
        name = filename
        ext = ''
    
    # Nettoyer le nom
    # Remplacer les caractères spéciaux par des underscores
    name = re.sub(r'[^\w\s-]', '_', name)
    # Remplacer les espaces par des underscores
    name = re.sub(r'[\s]+', '_', name)
    # Supprimer les underscores multiples
    name = re.sub(r'_+', '_', name)
    # Limiter la longueur
    name = name[:50]
    
    # Ajouter un timestamp pour l'unicité
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    return f"{name}_{timestamp}{ext}"


# ============================================================================
# PARSING DE DATES
# ============================================================================

def parse_date_fr(date_string: Any) -> date:
    """
    Parse une date au format français ou international.
    
    Args:
        date_string: String de date à parser (ou objet date/datetime)
        
    Returns:
        date: Date parsée
        
    Raises:
        ValidationError: Si la date ne peut pas être parsée
    """
    # Si c'est déjà une date
    if isinstance(date_string, date):
        return date_string
    
    if isinstance(date_string, datetime):
        return date_string.date()
    
    # Convertir en string et nettoyer
    if date_string is None:
        raise ValidationError("❌ Date manquante")
    
    date_string = str(date_string).strip()
    
    if not date_string:
        raise ValidationError("❌ Date vide")
    
    # Formats de date à essayer (ordre de priorité)
    formats = [
        '%d/%m/%Y',     # 31/12/2025 (français)
        '%d-%m-%Y',     # 31-12-2025
        '%d.%m.%Y',     # 31.12.2025
        '%Y-%m-%d',     # 2025-12-31 (ISO)
        '%Y/%m/%d',     # 2025/12/31
        '%d/%m/%y',     # 31/12/25 (année courte)
        '%d-%m-%y',     # 31-12-25
        '%d %B %Y',     # 31 décembre 2025
        '%d %b %Y',     # 31 déc 2025
        '%B %d, %Y',    # December 31, 2025
        '%b %d, %Y',    # Dec 31, 2025
    ]
    
    # Mapper les mois français
    french_months = {
        'janvier': 'January', 'février': 'February', 'mars': 'March',
        'avril': 'April', 'mai': 'May', 'juin': 'June',
        'juillet': 'July', 'août': 'August', 'septembre': 'September',
        'octobre': 'October', 'novembre': 'November', 'décembre': 'December',
        'janv': 'Jan', 'févr': 'Feb', 'mars': 'Mar',
        'avr': 'Apr', 'mai': 'May', 'juin': 'Jun',
        'juil': 'Jul', 'août': 'Aug', 'sept': 'Sep',
        'oct': 'Oct', 'nov': 'Nov', 'déc': 'Dec'
    }
    
    # Remplacer les mois français par anglais pour le parsing
    date_string_en = date_string
    for fr_month, en_month in french_months.items():
        date_string_en = date_string_en.replace(fr_month, en_month)
    
    # Essayer chaque format
    for fmt in formats:
        try:
            # Essayer avec la version française originale
            return datetime.strptime(date_string, fmt).date()
        except ValueError:
            try:
                # Essayer avec la version anglicisée
                return datetime.strptime(date_string_en, fmt).date()
            except ValueError:
                continue
    
    # Si aucun format ne fonctionne
    raise ValidationError(
        f"❌ Format de date non reconnu: '{date_string}'. "
        f"Utilisez le format JJ/MM/AAAA (ex: 31/12/2025)"
    )


# ============================================================================
# VALIDATION DE CHAMPS SPÉCIFIQUES
# ============================================================================

def validate_reference(reference: str) -> str:
    """
    Valide et formate une référence de projet AMDEC.
    
    Args:
        reference: Référence à valider
        
    Returns:
        str: Référence validée et formatée
        
    Raises:
        ValidationError: Si la référence est invalide
    """
    if not reference:
        # Générer une référence par défaut
        return f"AMDEC-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    
    reference = reference.strip().upper()
    
    # Vérifier le format (lettres, chiffres, tirets)
    if not re.match(r'^[A-Z0-9\-_]+$', reference):
        raise ValidationError(
            f"❌ Référence invalide: '{reference}'. "
            f"Utilisez uniquement lettres, chiffres et tirets."
        )
    
    # Vérifier la longueur
    if len(reference) > 50:
        raise ValidationError(
            f"❌ Référence trop longue: {len(reference)} caractères. "
            f"Maximum autorisé: 50 caractères."
        )
    
    if len(reference) < 3:
        raise ValidationError(
            f"❌ Référence trop courte: {len(reference)} caractères. "
            f"Minimum requis: 3 caractères."
        )
    
    return reference


def validate_team_members(team_members: str) -> str:
    """
    Valide et formate une liste de membres d'équipe.
    
    Args:
        team_members: String contenant les noms (séparés par virgules)
        
    Returns:
        str: Liste formatée des membres
    """
    if not team_members:
        return ''
    
    # Nettoyer et séparer
    members = [
        clean_text(member.strip(), max_length=100, preserve_newlines=False)
        for member in team_members.split(',')
    ]
    
    # Filtrer les entrées vides
    members = [m for m in members if m]
    
    # Limiter le nombre de membres
    if len(members) > 20:
        logger.warning(f"Trop de membres d'équipe: {len(members)}, limité à 20")
        members = members[:20]
    
    return ', '.join(members)


def validate_email_field(email: str) -> str:
    """
    Valide un champ email.
    
    Args:
        email: Email à valider
        
    Returns:
        str: Email validé
        
    Raises:
        ValidationError: Si l'email est invalide
    """
    if not email:
        return ''
    
    email = email.strip().lower()
    
    try:
        validate_email(email)
    except ValidationError:
        raise ValidationError(
            f"❌ Email invalide: '{email}'. "
            f"Format attendu: nom@domaine.com"
        )
    
    return email


# ============================================================================
# VALIDATION DE DONNÉES COMPLÈTES
# ============================================================================

def validate_failure_data(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Valide les données complètes d'une défaillance.
    
    Args:
        data: Dictionnaire des données de défaillance
        
    Returns:
        dict: Données validées et nettoyées
        
    Raises:
        ValidationError: Si les données sont invalides
    """
    errors = []
    cleaned = {}
    
    # Champs requis
    required_fields = ['component', 'failure_mode', 'gravity', 'occurrence', 'detection']
    
    for field in required_fields:
        if field not in data or not data[field]:
            errors.append(f"❌ Champ requis manquant: {field}")
    
    if errors:
        raise ValidationError(errors)
    
    # Nettoyer les champs texte
    text_fields = ['component', 'failure_mode', 'potential_cause', 'effect', 'preventive_actions']
    for field in text_fields:
        if field in data:
            cleaned[field] = clean_text(data[field], max_length=1000)
    
    # Valider les scores
    try:
        cleaned['gravity'], cleaned['occurrence'], cleaned['detection'] = validate_criticality_scores(
            data.get('gravity'),
            data.get('occurrence'),
            data.get('detection')
        )
    except ValidationError as e:
        errors.extend(e.messages if hasattr(e, 'messages') else [str(e)])
    
    # Champs optionnels
    if 'order' in data:
        try:
            cleaned['order'] = int(data['order'])
        except (ValueError, TypeError):
            cleaned['order'] = 0
    
    if 'status' in data:
        valid_statuses = ['PENDING', 'IN_PROGRESS', 'COMPLETED', 'CANCELLED']
        if data['status'] in valid_statuses:
            cleaned['status'] = data['status']
        else:
            errors.append(f"❌ Statut invalide: {data['status']}")
    
    if errors:
        raise ValidationError(errors)
    
    return cleaned


# ============================================================================
# HELPERS UTILITAIRES
# ============================================================================

def get_criticality_level(criticality: int) -> str:
    """
    Détermine le niveau de criticité basé sur le score.
    
    Args:
        criticality: Score de criticité (G × O × D)
        
    Returns:
        str: Niveau ('FAIBLE', 'MODÉRÉE', 'ÉLEVÉE')
    """
    if criticality <= 60:
        return "FAIBLE"
    elif criticality <= 120:
        return "MODÉRÉE"
    else:
        return "ÉLEVÉE"


def get_criticality_color(criticality: int) -> str:
    """
    Retourne la couleur associée au niveau de criticité.
    
    Args:
        criticality: Score de criticité
        
    Returns:
        str: Code couleur hexadécimal
    """
    level = get_criticality_level(criticality)
    colors = {
        "FAIBLE": "#27ae60",    # Vert
        "MODÉRÉE": "#f39c12",   # Orange
        "ÉLEVÉE": "#e74c3c"     # Rouge
    }
    return colors.get(level, "#95a5a6")  # Gris par défaut


def format_percentage(value: float, decimals: int = 1) -> str:
    """
    Formate une valeur en pourcentage.
    
    Args:
        value: Valeur entre 0 et 1
        decimals: Nombre de décimales
        
    Returns:
        str: Pourcentage formaté (ex: "75.5%")
    """
    if not isinstance(value, (int, float)):
        return "0%"
    
    percentage = value * 100
    return f"{percentage:.{decimals}f}%"


def validate_project_data(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Valide les données complètes d'un projet AMDEC.
    
    Args:
        data: Dictionnaire des données du projet
        
    Returns:
        dict: Données validées et nettoyées
        
    Raises:
        ValidationError: Si les données sont invalides
    """
    errors = []
    cleaned = {}
    
    # Champs requis
    if not data.get('name'):
        errors.append("❌ Nom du projet manquant")
    else:
        cleaned['name'] = clean_text(data['name'], max_length=200, preserve_newlines=False)
    
    # Référence
    try:
        cleaned['reference'] = validate_reference(data.get('reference', ''))
    except ValidationError as e:
        errors.append(str(e))
    
    # Client
    cleaned['client'] = clean_text(
        data.get('client', 'Non spécifié'),
        max_length=200,
        preserve_newlines=False
    )
    
    # Date d'analyse
    if 'analysis_date' in data:
        try:
            cleaned['analysis_date'] = parse_date_fr(data['analysis_date'])
        except ValidationError as e:
            errors.append(str(e))
    else:
        cleaned['analysis_date'] = date.today()
    
    # Équipe
    cleaned['team_members'] = validate_team_members(data.get('team_members', ''))
    
    # Objectif
    cleaned['objective'] = clean_text(data.get('objective', ''), max_length=1000)
    
    # Description
    cleaned['description'] = clean_text(data.get('description', ''), max_length=2000)
    
    if errors:
        raise ValidationError(errors)
    
    return cleaned
